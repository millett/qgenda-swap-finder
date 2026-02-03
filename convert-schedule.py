#!/usr/bin/env python3
"""
Convert QGenda schedule from XLSX to JavaScript format.

Usage:
    python convert-schedule.py [input_file.xlsx]

If no input file is specified, defaults to schedule.xlsx in the same directory.
"""

import sys
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# Color mapping for person types (approximate RGB values)
# Colors can be in ARGB format (00RRGGBB or FFRRGGBB)
# Verified mappings from user:
#   CA1 (purple): 009933FF (Koenig)
#   CA2 (cyan): 0000FFFF (Millett)
#   CA3 (teal): 0099CCCC (Yeker)
#   CRNA (light yellow-green): 00FFFF99 (Novelli)
COLOR_TO_TYPE = {
    # CA1 = Purple (009933FF = violet/purple)
    '009933FF': 'ca1',
    'FF9933FF': 'ca1',
    'FF800080': 'ca1',
    '00800080': 'ca1',
    'FF660066': 'ca1',
    'FF993399': 'ca1',
    'FFCC99FF': 'ca1',
    'FF7030A0': 'ca1',
    '007030A0': 'ca1',

    # CA2 = Cyan/Turquoise (0000FFFF = cyan)
    '0000FFFF': 'ca2',
    'FF00FFFF': 'ca2',
    '00CCFFFF': 'ca2',  # Light cyan - likely CA2
    'FFCCFFFF': 'ca2',
    'FF0000FF': 'ca2',
    '000000FF': 'ca2',
    'FF0066CC': 'ca2',
    'FF0070C0': 'ca2',
    '000070C0': 'ca2',
    'FF00B0F0': 'ca2',
    'FF4472C4': 'ca2',
    '004472C4': 'ca2',
    'FF5B9BD5': 'ca2',
    '005B9BD5': 'ca2',

    # CA3 = Teal/Grey-blue (0099CCCC = teal)
    '0099CCCC': 'ca3',
    'FF99CCCC': 'ca3',
    'FF808080': 'ca3',
    '00808080': 'ca3',
    'FF8EA9DB': 'ca3',
    '008EA9DB': 'ca3',
    'FF9BC2E6': 'ca3',
    'FFA6A6A6': 'ca3',
    'FFB4C6E7': 'ca3',
    '00B4C6E7': 'ca3',
    'FF8FAADC': 'ca3',
    '008FAADC': 'ca3',

    # Orange = Fellow
    'FFFF6600': 'fellow',
    '00FF6600': 'fellow',
    'FFFF9900': 'fellow',
    'FFFFC000': 'fellow',
    '00FFC000': 'fellow',
    'FFED7D31': 'fellow',
    '00ED7D31': 'fellow',
    'FFF4B084': 'fellow',

    # CRNA = Light yellow-green (00FFFF99)
    '00FFFF99': 'crna',
    'FFFFFF99': 'crna',
    'FFFFFF00': 'crna',
    '00FFFF00': 'crna',
    'FFFFCC00': 'crna',
    'FFFFEB9C': 'crna',
    'FFFFF2CC': 'crna',
    '00FFF2CC': 'crna',
    'FFFFD966': 'crna',
    '00FFD966': 'crna',

    # Red = announcement/other
    'FFFF0000': 'other',
    '00FF0000': 'other',

    # Pink/Salmon = faculty
    '00FFCCCC': 'faculty',
    'FFFFCCCC': 'faculty',

    # Light lavender/periwinkle = faculty (Awad, Patel, etc.)
    '00CCCCFF': 'faculty',
    'FFCCCCFF': 'faculty',
    '0099FFFF': 'faculty',
    'FF99FFFF': 'faculty',

    # Light green = unknown, classify as other for now
    '0099FF99': 'other',
}


def get_cell_color(cell):
    """Extract the background color from a cell."""
    try:
        fill = cell.fill
        if fill and fill.patternType == 'solid':
            fg_color = fill.fgColor
            if fg_color and fg_color.rgb:
                return fg_color.rgb
    except:
        pass
    return None


def classify_by_color(color_rgb):
    """Classify person type by cell color."""
    if not color_rgb:
        return None

    # Direct match
    if color_rgb in COLOR_TO_TYPE:
        return COLOR_TO_TYPE[color_rgb]

    # Try to match by color similarity (simplified)
    try:
        if len(color_rgb) >= 6:
            # Extract RGB values (handle ARGB format - skip first 2 chars if 8 chars)
            rgb = color_rgb
            if len(rgb) == 8:
                rgb = rgb[2:]  # Skip alpha

            r = int(rgb[0:2], 16)
            g = int(rgb[2:4], 16)
            b = int(rgb[4:6], 16)

            # Purple/Violet: medium-high red, low green, high blue
            if 80 < r < 200 and g < 80 and b > 200:
                return 'ca1'

            # Blue: low red, any green, high blue
            if r < 100 and b > 150:
                return 'ca2'

            # Grey-blue: medium everything with blue tint
            if 100 < r < 200 and 130 < g < 220 and 180 < b < 240:
                return 'ca3'

            # Orange: high red, medium green, low blue
            if r > 200 and 80 < g < 200 and b < 130:
                return 'fellow'

            # Yellow: high red, high green, low-medium blue
            if r > 200 and g > 200 and b < 180:
                return 'crna'

            # Grey: similar R, G, B values
            if abs(r - g) < 30 and abs(g - b) < 30 and abs(r - b) < 30:
                if r > 150:  # Light grey
                    return 'ca3'
    except:
        pass

    return None


def parse_qgenda_excel(file_path: Path) -> tuple[pd.DataFrame, dict]:
    """Parse QGenda Excel export into a clean DataFrame and extract person colors."""
    # Load workbook with openpyxl to get colors
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    # Also load with pandas for easier data handling
    df = pd.read_excel(file_path, header=None, engine='openpyxl')

    records = []
    current_dates = {}
    person_colors = {}  # name -> color RGB

    for idx, row in df.iterrows():
        excel_row = idx + 1  # Excel rows are 1-indexed

        # Check if this is a date header row
        try:
            if pd.notna(row[0]) and '202' in str(row[0]):
                for day_idx, col in enumerate([0, 2, 4, 6, 8, 10, 12]):
                    if pd.notna(row[col]):
                        try:
                            current_dates[day_idx] = pd.to_datetime(row[col])
                        except:
                            pass
                continue
        except:
            pass

        # Process data rows
        for day_idx, (name_col, shift_col) in enumerate([(0,1), (2,3), (4,5), (6,7), (8,9), (10,11), (12,13)]):
            if day_idx in current_dates and pd.notna(row[name_col]) and pd.notna(row[shift_col]):
                name = str(row[name_col]).strip()
                shift = str(row[shift_col]).strip()
                skip_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday', 'Printed']
                if name and shift and not any(s in name for s in skip_names):
                    records.append({
                        'date': current_dates[day_idx],
                        'name': name,
                        'shift': shift
                    })

                    # Extract color for person classification
                    if name not in person_colors:
                        try:
                            cell = ws.cell(row=excel_row, column=name_col + 1)  # openpyxl is 1-indexed
                            color = get_cell_color(cell)
                            if color:
                                person_colors[name] = color
                        except:
                            pass

    wb.close()
    return pd.DataFrame(records), person_colors


def convert_to_javascript(df: pd.DataFrame, person_colors: dict, output_path: Path, input_file: Path):
    """Convert DataFrame to JavaScript format and write to file."""
    # Include CA, CRNA, Faculty, and Fellow shifts for proper classification
    valid_prefixes = ('CA ', 'CRNA', 'Faculty', 'Fellow')
    filtered_shifts = df[df['shift'].str.startswith(valid_prefixes, na=False)].copy()

    # Sort by date
    filtered_shifts = filtered_shifts.sort_values('date')

    # Classify people by color
    person_types = {}
    unknown_colors = {}
    for name, color in person_colors.items():
        ptype = classify_by_color(color)
        if ptype:
            person_types[name] = ptype
        else:
            unknown_colors[name] = color

    # Also classify by shift prefix as fallback
    for name in filtered_shifts['name'].unique():
        if name in person_types:
            continue
        person_shifts = filtered_shifts[filtered_shifts['name'] == name]['shift'].tolist()
        has_ca = any(s.startswith('CA ') for s in person_shifts)
        has_crna = any('CRNA' in s for s in person_shifts)
        has_faculty = any(s.startswith('Faculty') for s in person_shifts)
        has_fellow = any(s.startswith('Fellow') for s in person_shifts)

        if has_crna and not has_ca:
            person_types[name] = 'crna'
        elif has_faculty:
            person_types[name] = 'faculty'
        elif has_fellow:
            person_types[name] = 'fellow'
        elif has_ca:
            person_types[name] = 'resident'  # Unknown year level

    # Generate timestamp
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # Build JavaScript content
    js_lines = [
        f"// Generated from {input_file.name}",
        f"// Last updated: {timestamp}",
        "const SCHEDULE = ["
    ]

    # Add data rows
    for _, row in filtered_shifts.iterrows():
        date_str = row['date'].strftime('%Y-%m-%d')
        name = row['name'].replace('"', '\\"')  # Escape quotes
        shift = row['shift'].replace('"', '\\"')
        js_lines.append(f'  {{ date: "{date_str}", name: "{name}", shift: "{shift}" }},')

    js_lines.append("];")
    js_lines.append("")

    # Add person types mapping
    js_lines.append("// Person types: ca1, ca2, ca3, fellow, crna, faculty, resident (unknown year)")
    js_lines.append("const PERSON_TYPES_DATA = {")
    for name, ptype in sorted(person_types.items()):
        escaped_name = name.replace('"', '\\"')
        js_lines.append(f'  "{escaped_name}": "{ptype}",')
    js_lines.append("};")

    # Write to file
    output_path.write_text('\n'.join(js_lines))

    # Print color analysis for debugging
    if unknown_colors:
        print(f"\nUnknown colors found ({len(unknown_colors)} people):")
        for name, color in list(unknown_colors.items())[:10]:
            print(f"  {name}: {color}")

    print(f"\nPerson type breakdown:")
    type_counts = {}
    for ptype in person_types.values():
        type_counts[ptype] = type_counts.get(ptype, 0) + 1
    for ptype, count in sorted(type_counts.items()):
        print(f"  {ptype}: {count}")

    return len(filtered_shifts)


def main():
    # Get script directory
    script_dir = Path(__file__).parent

    # Determine input file
    if len(sys.argv) > 1:
        input_file = Path(sys.argv[1])
    else:
        input_file = script_dir / "schedule.xlsx"

    # Check if input file exists
    if not input_file.exists():
        print(f"Error: Input file not found: {input_file}")
        sys.exit(1)

    # Output file path
    output_file = script_dir / "js" / "schedule.js"
    output_file.parent.mkdir(exist_ok=True)

    print(f"Reading schedule from: {input_file}")

    # Parse Excel file
    try:
        df, person_colors = parse_qgenda_excel(input_file)
        print(f"Parsed {len(df)} total schedule entries")
        print(f"Found colors for {len(person_colors)} people")
    except Exception as e:
        print(f"Error parsing Excel file: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

    # Convert to JavaScript
    try:
        num_records = convert_to_javascript(df, person_colors, output_file, input_file)
        print(f"Exported {num_records} shift records to: {output_file}")
        print(f"✓ Successfully generated {output_file.name}")
    except Exception as e:
        print(f"Error generating JavaScript file: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
